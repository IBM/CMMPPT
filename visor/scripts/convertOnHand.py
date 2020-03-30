# This python script requires the openpyxl python library
# This script converts the onHand.xlsx spreadsheet to a CSV file
# This script requires two parameters
# (1) the name of the spreadsheet to be converted --- ex: fname.xlsx
# (2) the name of the CSV output file produced --- ex: foutput.csv

import openpyxl
import csv
import datetime
import sys

fmt = '%Y-%m-%d %H:%M:%S'

theSpreadSheet = sys.argv[1]
theOutputCSV = sys.argv[2]

wb = openpyxl.load_workbook(theSpreadSheet)
sheet = wb.active
with open(theOutputCSV, 'w', newline="") as f:
    c = csv.writer(f,quoting=csv.QUOTE_NONNUMERIC)
    last_row = sheet.max_row
    while sheet.cell(column=1, row=last_row).value is None and last_row > 0:
        last_row -= 1
    last_col_a_value = sheet.cell(column=1, row=last_row).value
    myHeader = []
    for r in sheet.iter_rows(min_row=1, min_col=1, max_row=last_row, max_col=5):
        myList = []
        for idx, cell in enumerate(r):
            myList.append(cell.value)
        c.writerow(myList)
f.close()
